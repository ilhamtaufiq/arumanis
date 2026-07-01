#!/usr/bin/env python3
"""Convert daftar penerima DOCX/PDF sources into import-ready Excel templates."""

from __future__ import annotations

import json
import re
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

DOWNLOADS = Path(r"C:\Users\asusg\Downloads")
DOCX_KUBANG = DOWNLOADS / "DAFTAR PENERIMA AIR BERSIH DESA KUBANG KEC. PASIR KUDA (1).docx"
RAWABELUT_JSON = DOWNLOADS / "rawabelut_parsed.json"

FOTO_LEVELS = ("0", "25", "50", "75", "100")

HEADERS = [
    "no",
    "nama",
    "nik",
    "alamat",
    "jumlah_jiwa",
    "latitude",
    "longitude",
    *[f"nama_file_foto_{level}" for level in FOTO_LEVELS],
]

W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
R_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
A_NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}"


def foto_name_for_row(row_no: int, level: str = "0") -> str:
    return f"{row_no:03d}_{level}.jpg"


def normalize_legacy_foto_name(raw: str, row_no: int) -> str:
    text = str(raw or "").strip()
    if not text:
        return ""
    if re.search(r"_\d+\.jpe?g$", text, re.I):
        return text
    match = re.match(r"^(\d+)\.jpe?g$", text, re.I)
    if match:
        return foto_name_for_row(int(match.group(1)), "0")
    return foto_name_for_row(row_no, "0")


def row_to_excel_values(row: dict) -> list:
    row_no = int(row.get("no") or 0)
    legacy_foto = row.get("nama_file_foto_0") or row.get("nama_file_foto") or ""
    foto_0 = normalize_legacy_foto_name(legacy_foto, row_no)

    return [
        row.get("no"),
        row.get("nama"),
        row.get("nik"),
        row.get("alamat"),
        row.get("jumlah_jiwa"),
        row.get("latitude"),
        row.get("longitude"),
        foto_0,
        row.get("nama_file_foto_25") or "",
        row.get("nama_file_foto_50") or "",
        row.get("nama_file_foto_75") or "",
        row.get("nama_file_foto_100") or "",
    ]


def clean_coord(raw: str, is_longitude: bool = False) -> float | None:
    if not raw:
        return None
    text = re.sub(r"\s+", "", raw).replace(",", ".")
    text = re.sub(r"\.+", ".", text)
    try:
        value = float(text)
        if abs(value) <= (180 if is_longitude else 90):
            if is_longitude and 0 < value < 20:
                digits = re.sub(r"\D", "", raw)
                if len(digits) >= 8:
                    value = float(f"{digits[:3]}.{digits[3:]}")
            return round(value, 6)
    except ValueError:
        pass

    digits = re.sub(r"\D", "", raw)
    if not digits:
        return None
    sign = -1 if "-" in raw else 1
    if is_longitude and len(digits) >= 8:
        return round(sign * float(f"{digits[:3]}.{digits[3:]}"), 6)
    if not is_longitude and len(digits) >= 6:
        for cut in (1, 2):
            candidate = sign * float(f"{digits[:cut]}.{digits[cut:]}")
            if abs(candidate) <= 90:
                return round(candidate, 6)
    return None


def split_nama_alamat(raw: str) -> tuple[str, str]:
    text = re.sub(r"\s+", " ", raw).strip()
    match = re.match(r"^(.+?)(?=(?:Kp\.?|KP\.?|RT\.?|RW\.?|Ds\.?|Desa\s))", text, re.I)
    if match:
        return match.group(1).strip(), text[match.end() :].strip()
    return text, ""


def parse_nik(raw: str) -> str:
    return re.sub(r"\D", "", raw)[:16]


def parse_jiwa(raw: str) -> int:
    match = re.search(r"(\d+)", raw)
    return int(match.group(1)) if match else 1


def cell_text(cell) -> str:
    parts = []
    for node in cell.iter(f"{W_NS}t"):
        if node.text:
            parts.append(node.text)
        if node.tail:
            parts.append(node.tail)
    return "".join(parts).strip()


def row_images(row, rid_to_target: dict[str, str]) -> list[str]:
    images = []
    for blip in row.iter(f"{A_NS}blip"):
        embed = blip.attrib.get(f"{R_NS}embed")
        if embed and embed in rid_to_target:
            images.append(rid_to_target[embed])
    return images


def parse_kubang_docx() -> list[dict]:
    import xml.etree.ElementTree as ET

    rows: list[dict] = []
    foto_dir = DOWNLOADS / "foto_penerima_kubang"
    if foto_dir.exists():
        for old_file in foto_dir.glob("*.jpg"):
            old_file.unlink()
    foto_dir.mkdir(exist_ok=True)

    with zipfile.ZipFile(DOCX_KUBANG) as archive:
        root = ET.fromstring(archive.read("word/document.xml"))
        rels = ET.fromstring(archive.read("word/_rels/document.xml.rels"))
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels
            if "Id" in rel.attrib and "Target" in rel.attrib
        }

        image_index = 0
        for table_row in root.iter(f"{W_NS}tr"):
            cells = [cell_text(cell) for cell in table_row.iter(f"{W_NS}tc")]
            joined = " ".join(cells).upper()
            if not cells or ("NAMA" in joined and "JIWA" in joined):
                continue

            padded = cells + [""] * 6
            no_raw, nama_alamat, jiwa_raw, koord_raw, _, nik_raw = padded[:6]
            if not nama_alamat.strip():
                continue

            nama, alamat = split_nama_alamat(nama_alamat)
            lat_match = re.search(r"lat[:\s]*(-?[\d.,\s]+)", koord_raw, re.I)
            lng_match = re.search(r"long[:\s]*(-?[\d.,\s]+)", koord_raw, re.I)
            images = row_images(table_row, rid_to_target)

            no_match = re.search(r"(\d+)", no_raw)
            row_no = int(no_match.group(1)) if no_match else len(rows) + 1
            foto_name = ""

            if images:
                image_index += 1
                media_path = images[-1]
                zip_path = media_path if media_path.startswith("word/") else f"word/{media_path}"
                foto_name = foto_name_for_row(row_no, "0")
                (foto_dir / foto_name).write_bytes(archive.read(zip_path))

            latitude = clean_coord(lat_match.group(1), False) if lat_match else None
            longitude = clean_coord(lng_match.group(1), True) if lng_match else None
            if longitude is not None and 0 < longitude < 20 and latitude and abs(latitude) > 5:
                digits = re.sub(r"\D", "", lng_match.group(1) if lng_match else "")
                if len(digits) >= 8:
                    longitude = round(float(f"{digits[:3]}.{digits[3:]}"), 6)

            rows.append(
                {
                    "no": row_no,
                    "nama": nama,
                    "nik": parse_nik(nik_raw),
                    "alamat": alamat,
                    "jumlah_jiwa": parse_jiwa(jiwa_raw),
                    "latitude": latitude,
                    "longitude": longitude,
                    "nama_file_foto_0": foto_name,
                }
            )

    zip_output = DOWNLOADS / "foto_penerima_kubang.zip"
    with zipfile.ZipFile(zip_output, "w", zipfile.ZIP_DEFLATED) as foto_zip:
        for image_file in sorted(foto_dir.glob("*.jpg")):
            foto_zip.write(image_file, arcname=image_file.name)

    return rows


def write_workbook(rows: list[dict], output_path: Path, source_title: str, zip_name: str | None = None) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"

    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(row_to_excel_values(row))

    widths = [5, 24, 18, 42, 12, 14, 14, 16, 16, 16, 16, 16]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    guide = workbook.create_sheet("Petunjuk")
    guide.append([f"Sumber: {source_title}"])
    guide.append([])
    guide.append(["Template import Arumanis — format foto progress 0%, 25%, 50%, 75%, 100%."])
    guide.append(["Koordinat satu baris dipakai untuk semua level foto pada baris yang sama."])
    guide.append([])
    guide.append(["Kolom foto:", ", ".join(f"nama_file_foto_{level}" for level in FOTO_LEVELS)])
    guide.append(["Contoh nama file ZIP:", "001_0.jpg, 001_25.jpg, 001_50.jpg, 001_75.jpg, 001_100.jpg"])
    if zip_name:
        guide.append([f"ZIP foto: {zip_name}"])
    guide.append([])
    guide.append(["Saat import di aplikasi, pilih komponen output yang sesuai lalu unggah Excel + ZIP."])
    guide.column_dimensions["A"].width = 90

    workbook.save(output_path)


def main() -> None:
    kubang_rows = parse_kubang_docx()
    kubang_output = DOWNLOADS / "template_import_penerima_KUBANG.xlsx"
    write_workbook(
        kubang_rows,
        kubang_output,
        "DAFTAR PENERIMA AIR BERSIH DESA KUBANG KEC. PASIR KUDA",
        zip_name="foto_penerima_kubang.zip",
    )
    foto_count = sum(1 for row in kubang_rows if row.get("nama_file_foto_0"))
    print(f"KUBANG: {len(kubang_rows)} baris, {foto_count} foto 0% -> {kubang_output}")

    if RAWABELUT_JSON.exists():
        rawabelut_rows = json.loads(RAWABELUT_JSON.read_text(encoding="utf-8"))
        rawabelut_output = DOWNLOADS / "template_import_penerima_RAWABELUT.xlsx"
        write_workbook(
            rawabelut_rows,
            rawabelut_output,
            "DAFTAR NAMA WARGA PENERIMA PAM DESA RAWABELUT",
        )
        print(f"RAWABELUT: {len(rawabelut_rows)} baris -> {rawabelut_output}")


if __name__ == "__main__":
    main()