#!/usr/bin/env python3
"""Generate FLOWCHART SOP Excel (format LAMPIRAN / tabel + alur)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "docs" / "SOP_PENGGUNAAN_ARUMANIS.xlsx"

# Colors
ORANGE = PatternFill("solid", fgColor="F4B183")
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FILL = PatternFill("solid", fgColor="BDD7EE")
FLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
DECISION_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE = PatternFill("solid", fgColor="FFFFFF")
MARK = "●"

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def sty(cell, *, bold=False, size=10, fill=None, wrap=True, center=False, h="left"):
    cell.font = Font(name="Arial", size=size, bold=bold)
    cell.border = BORDER
    cell.alignment = Alignment(
        horizontal="center" if center else h,
        vertical="center",
        wrap_text=wrap,
    )
    if fill:
        cell.fill = fill


def write_header(ws, title: str, lampiran: str = "LAMPIRAN") -> int:
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = lampiran
    sty(c, bold=True, size=11, center=True)

    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = (
        "Standar Operasional Prosedur (SOP) Penggunaan Aplikasi ARUMANIS "
        "dan Panel Pengawasan — Satu Data Air Minum dan Sanitasi Kabupaten Cianjur"
    )
    sty(c, size=9, center=True, wrap=True)
    ws.row_dimensions[2].height = 36

    ws.merge_cells("A3:K3")
    sty(ws["A3"], size=9, center=True)
    ws["A3"].value = "Nomor : ....................................          Tanggal : 1 Juli 2026"

    ws.merge_cells("A4:K4")
    c = ws["A4"]
    c.value = title
    sty(c, bold=True, size=12, fill=TITLE_FILL, center=True)
    ws.row_dimensions[4].height = 28

    headers = [
        "No",
        "Kegiatan",
        "Flowchart",
        "Admin",
        "Operator",
        "Pengawas",
        "Sistem",
        "Kategori",
        "Waktu",
        "Output",
        "Ket",
    ]
    row = 5
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        sty(cell, bold=True, size=9, fill=HEADER_FILL, center=True)
    ws.row_dimensions[row].height = 32
    return row + 1


def add_step(
    ws,
    row: int,
    no: str,
    kegiatan: str,
    flow: str,
    roles: dict[str, bool],
    kategori: str,
    waktu: str,
    output: str,
    ket: str = "",
    flow_fill=None,
    height: int = 72,
) -> int:
    data = [
        no,
        kegiatan,
        flow,
        MARK if roles.get("admin") else "",
        MARK if roles.get("operator") else "",
        MARK if roles.get("pengawas") else "",
        MARK if roles.get("sistem") else "",
        kategori,
        waktu,
        output,
        ket,
    ]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        fill = None
        if col == 3:
            fill = flow_fill or FLOW_FILL
        elif col in (4, 5, 6, 7) and val == MARK:
            fill = ORANGE
        sty(cell, size=9, fill=fill, center=(col in (1, 4, 5, 6, 7, 8, 9)))
    ws.row_dimensions[row].height = height
    return row + 1


def setup_columns(ws):
    widths = {
        "A": 5,
        "B": 42,
        "C": 22,
        "D": 8,
        "E": 9,
        "F": 9,
        "G": 8,
        "H": 14,
        "I": 10,
        "J": 22,
        "K": 18,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def footer_note(ws, row: int, text: str):
    ws.merge_cells(f"A{row}:K{row}")
    c = ws.cell(row=row, column=1, value=text)
    sty(c, size=8, center=True)


def sheet_login(wb: Workbook):
    ws = wb.create_sheet("SOP-01 Login & SSO", 0)
    setup_columns(ws)
    r = write_header(ws, "FLOWCHART SOP AKSES & AUTENTIKASI ARUMANIS / PANEL PENGAWASAN")

    r = add_step(
        ws,
        r,
        "1",
        "Pengguna membuka URL aplikasi Arumanis di browser (Chrome/Firefox/Edge terbaru).",
        "┌─────────┐\n│  User   │\n└────┬────┘\n     ↓",
        {"admin": True, "operator": True, "pengawas": True},
        "Akses",
        "± 2 menit",
        "Halaman /sign-in tampil",
    )
    r = add_step(
        ws,
        r,
        "2",
        "Pengguna memasukkan email dan password akun APIAMIS, lalu klik Sign In "
        "(atau Google Login jika diaktifkan).",
        "┌──────────┐\n│  Login   │\n└────┬─────┘\n     ↓",
        {"admin": True, "operator": True, "pengawas": True},
        "Autentikasi",
        "± 1 menit",
        "Sesi cookie httpOnly terbentuk",
    )
    r = add_step(
        ws,
        r,
        "3",
        "Sistem memvalidasi kredensial melalui BFF Arumanis ke APIAMIS.\n"
        "Jika gagal → tampilkan pesan error, kembali ke langkah 2.",
        "   ◇ Berhasil?\n  ╱        ╲\nTidak      Ya\n ↓          ↓",
        {"sistem": True},
        "Validasi",
        "Real-time",
        "Token sesi valid / error",
        flow_fill=DECISION_FILL,
        height=88,
    )
    r = add_step(
        ws,
        r,
        "4",
        "Sistem mengecek peran (role) pengguna:\n"
        "• Admin / Operator / Viewer → Dashboard Arumanis (/dashboard)\n"
        "• Pengawas / Konsultan Pengawas → lanjut SSO ke Panel Pengawasan",
        "┌──────────┐\n│ Cek Role │\n└────┬─────┘\n     ↓",
        {"sistem": True},
        "Routing",
        "Real-time",
        "Halaman tujuan ditentukan",
    )
    r = add_step(
        ws,
        r,
        "5",
        "Untuk peran pengawas: sistem mengarahkan ke /pengawasan/login?token=... "
        "dan menyinkronkan token SSO → cookie pengawas_session.",
        "┌─────────┐\n│   SSO   │\n└────┬────┘\n     ↓",
        {"pengawas": True, "sistem": True},
        "SSO",
        "± 30 detik",
        "Dashboard Pengawasan (/pengawasan/)",
        ket="Tidak ada form login terpisah di panel",
    )
    r = add_step(
        ws,
        r,
        "6",
        "Pengguna menjalankan tugas sesuai SOP modul. "
        "Logout: klik avatar → Logout → kembali ke /sign-in.",
        "┌──────────┐\n│ Operasi  │\n└────┬─────┘\n     ↓\n┌──────────┐\n│  Logout  │\n└──────────┘",
        {"admin": True, "operator": True, "pengawas": True},
        "Operasional",
        "Sesuai tugas",
        "Sesi berakhir aman",
    )
    footer_note(ws, r + 1, "*** Flowchart SOP Akses — Arumanis v0.5.0 · Panel /pengawasan ***")


def sheet_input_program(wb: Workbook):
    ws = wb.create_sheet("SOP-02 Input Program")
    setup_columns(ws)
    r = write_header(ws, "FLOWCHART SOP INPUT PROGRAM BARU (ARUMANIS UTAMA)")

    steps = [
        (
            "1",
            "Operator/Admin login ke Arumanis dan membuka modul Program Kegiatan (/kegiatan).",
            "┌──────────┐\n│ Kegiatan │\n└────┬─────┘\n     ↓",
            {"admin": True, "operator": True},
            "Perencanaan",
            "± 15 menit",
            "Master kegiatan tersimpan",
            "",
            None,
            68,
        ),
        (
            "2",
            "Tambah kegiatan baru: isi nama, kode unik, sumber dana, pagu, tahun anggaran → Simpan.",
            "┌──────────┐\n│  Tambah  │\n└────┬─────┘\n     ↓",
            {"operator": True},
            "Input data",
            "± 10 menit",
            "Kegiatan muncul di daftar",
            "",
            None,
            68,
        ),
        (
            "3",
            "Buka modul Pekerjaan (/pekerjaan) → Tambah: pilih kegiatan, kecamatan, desa, "
            "nama paket, pagu → Simpan.",
            "┌──────────┐\n│ Pekerjaan│\n└────┬─────┘\n     ↓",
            {"operator": True},
            "Input data",
            "± 15 menit",
            "Pekerjaan terhubung kegiatan",
            "",
            None,
            72,
        ),
        (
            "4",
            "Tambah Output (/output) dan Penerima (/penerima) per pekerjaan sesuai RAB/kontrak.",
            "┌──────────┐\n│ Output & │\n│ Penerima │\n└────┬─────┘\n     ↓",
            {"operator": True},
            "Input data",
            "± 20 menit",
            "Komponen & penerima tercatat",
            "",
            None,
            80,
        ),
        (
            "5",
            "Upload dokumentasi awal di Berkas (/berkas) dan Foto (/foto) bila tersedia.",
            "┌──────────┐\n│  Berkas  │\n│   Foto   │\n└────┬─────┘\n     ↓",
            {"operator": True},
            "Dokumentasi",
            "± 15 menit",
            "File terunggah & tervalidasi",
            "Geo-fencing aktif untuk foto",
            None,
            80,
        ),
        (
            "6",
            "Verifikasi di Dashboard: metrik pekerjaan, output, dan data quality stats terupdate.",
            "   ◇ Lengkap?\n  ╱        ╲\nTidak      Ya\n ↓          ↓\nPerbaiki   Selesai",
            {"admin": True, "operator": True},
            "Verifikasi",
            "± 5 menit",
            "Data siap operasional",
            "Ulangi langkah yang kurang jika Tidak",
            DECISION_FILL,
            96,
        ),
    ]
    for s in steps:
        r = add_step(ws, r, *s)
    footer_note(ws, r + 1, "*** SOP Input Program — www/bun (Arumanis Utama) ***")


def sheet_kontrak(wb: Workbook):
    ws = wb.create_sheet("SOP-03 Kontrak")
    setup_columns(ws)
    r = write_header(ws, "FLOWCHART SOP PENGELOLAAN KONTRAK & PENYEDIA")

    steps = [
        (
            "1",
            "Pastikan master Pekerjaan dan Penyedia (/penyedia) sudah terdaftar.",
            "┌──────────┐\n│ Prasyarat│\n└────┬─────┘\n     ↓",
            {"operator": True},
            "Persiapan",
            "± 10 menit",
            "Data induk siap",
            "",
            None,
            68,
        ),
        (
            "2",
            "Buka Kontrak (/kontrak) → Tambah: pilih pekerjaan & penyedia, isi nomor, "
            "nilai, tanggal mulai/selesai → Simpan.",
            "┌──────────┐\n│  Kontrak │\n└────┬─────┘\n     ↓",
            {"operator": True},
            "Kontrak",
            "± 20 menit",
            "Kontrak aktif",
            "",
            None,
            72,
        ),
        (
            "3",
            "Jika ada perubahan kontrak: buat Addendum (/kontrak-addendums) "
            "dan update register dokumen (/pekerjaan/register).",
            "   ◇ Ada addendum?\n  ╱            ╲\nTidak          Ya\n ↓              ↓",
            {"operator": True},
            "Addendum",
            "Sesuai kebutuhan",
            "Nilai & dokumen sinkron",
            "",
            DECISION_FILL,
            88,
        ),
        (
            "4",
            "Admin/Operator verifikasi status kontrak di Dashboard dan modul terkait.",
            "┌──────────┐\n│ Verifikasi│\n└──────────┘",
            {"admin": True, "operator": True},
            "Monitoring",
            "± 5 menit",
            "Kontrak siap ditugaskan ke pengawas",
            "",
            None,
            68,
        ),
    ]
    for s in steps:
        r = add_step(ws, r, *s)
    footer_note(ws, r + 1, "*** SOP Kontrak — Arumanis Utama ***")


def sheet_pengawasan(wb: Workbook):
    ws = wb.create_sheet("SOP-04 Pengawasan Lapangan")
    setup_columns(ws)
    r = write_header(ws, "FLOWCHART SOP PEMANTAUAN & PELAPORAN LAPANGAN (PANEL PENGAWASAN)")

    steps = [
        (
            "1",
            "Pengawas login via SSO dari Arumanis → masuk Dashboard Pengawasan (/pengawasan/).",
            "┌──────────┐\n│   SSO    │\n└────┬─────┘\n     ↓",
            {"pengawas": True, "sistem": True},
            "Akses",
            "± 2 menit",
            "Dashboard KPI tampil",
            "",
            None,
            72,
        ),
        (
            "2",
            "Baca 4 KPI: jumlah paket, belum isi progress, deviasi, foto belum lengkap. "
            "Buka paket di section 'perlu perhatian'.",
            "┌──────────┐\n│Dashboard │\n└────┬─────┘\n     ↓",
            {"pengawas": True},
            "Pemantauan",
            "Harian",
            "Daftar prioritas paket",
            "",
            None,
            80,
        ),
        (
            "3",
            "Buka detail pekerjaan → tab Penerima: tambah/edit penerima manfaat "
            "(individu/komunal, NIK, jiwa).",
            "┌──────────┐\n│ Penerima │\n└────┬─────┘\n     ↓",
            {"pengawas": True},
            "Data lapangan",
            "Per paket",
            "Penerima lengkap",
            "",
            None,
            80,
        ),
        (
            "4",
            "Tab Foto: unggah foto per slot progress (0%, 25%, 50%, 75%, 100%) "
            "dengan koordinat GPS pada setiap output.",
            "┌──────────┐\n│   Foto   │\n└────┬─────┘\n     ↓",
            {"pengawas": True},
            "Dokumentasi",
            "Per kunjungan",
            "Matriks foto terisi",
            "Status: Belum ada / Belum Selesai / Selesai",
            None,
            88,
        ),
        (
            "5",
            "Tab Progress / Buat Laporan: pilih minggu aktif, isi Rencana & Realisasi "
            "per item → Simpan.",
            "┌──────────┐\n│ Progress │\n└────┬─────┘\n     ↓",
            {"pengawas": True},
            "Pelaporan",
            "Mingguan",
            "Progress & deviasi terupdate",
            "Sebelum Jumat",
            None,
            80,
        ),
        (
            "6",
            "Jika ada kendala teknis: buat Tiket di /pengawasan/tiket (kategori Permasalahan Lapangan). "
            "Jika tidak ada kendala, selesai siklus mingguan.",
            "   ◇ Kendala?\n  ╱        ╲\nTidak      Ya\n ↓          ↓\nSelesai    Buat Tiket",
            {"pengawas": True},
            "Tiket",
            "Sesuai kejadian",
            "Tiket tercatat & ditindaklanjuti",
            "Eskalasi ke admin via /tiket Arumanis",
            DECISION_FILL,
            96,
        ),
        (
            "7",
            "Admin/Operator di Arumanis memantau progress & tiket; data tersinkron dua arah via APIAMIS.",
            "┌──────────┐\n│  Sinkron │\n└──────────┘",
            {"admin": True, "operator": True, "sistem": True},
            "Integrasi",
            "Real-time",
            "Data kantor = data lapangan",
            "www/pengawas → APIAMIS → www/bun",
            None,
            72,
        ),
    ]
    for s in steps:
        r = add_step(ws, r, *s)
    footer_note(ws, r + 1, "*** SOP Panel Pengawasan — www/pengawas · /pengawasan ***")


def sheet_penugasan(wb: Workbook):
    ws = wb.create_sheet("SOP-05 Penugasan Pengawas")
    setup_columns(ws)
    r = write_header(ws, "FLOWCHART SOP PENUGASAN PENGAWAS (ADMIN → PANEL PENGAWASAN)")

    steps = [
        (
            "1",
            "Admin membuka master Pengawas (/pengawas): pastikan nama, NIP, "
            "dan data kontak benar.",
            "┌──────────┐\n│ Pengawas │\n└────┬─────┘\n     ↓",
            {"admin": True},
            "Master data",
            "± 10 menit",
            "Data pengawas valid",
            "",
            None,
            72,
        ),
        (
            "2",
            "Buka Users (/users): pastikan akun pengawas memiliki role "
            "pengawas atau konsultan_pengawas.",
            "┌──────────┐\n│   User   │\n└────┬─────┘\n     ↓",
            {"admin": True},
            "Akses",
            "± 5 menit",
            "Role pengawas aktif",
            "",
            None,
            68,
        ),
        (
            "3",
            "Assign Pekerjaan (/user-pekerjaan): tautkan user pengawas "
            "ke paket pekerjaan yang diawasi.",
            "┌──────────┐\n│  Assign  │\n└────┬─────┘\n     ↓",
            {"admin": True},
            "Penugasan",
            "± 15 menit",
            "Paket muncul di dashboard pengawas",
            "Wajib sebelum pengawas bekerja",
            None,
            80,
        ),
        (
            "4",
            "Pengawas login → verifikasi paket tampil di /pengawasan/pekerjaan.",
            "   ◇ Paket tampil?\n  ╱            ╲\nTidak          Ya\n ↓              ↓\nUlangi assign  Lanjut SOP-04",
            {"pengawas": True, "admin": True},
            "Verifikasi",
            "± 5 menit",
            "Penugasan efektif",
            "",
            DECISION_FILL,
            96,
        ),
        (
            "5",
            "(Opsional) Admin impersonate pengawas dari /users untuk review UX lapangan.",
            "┌──────────┐\n│Impersonate│\n└────┬──────┘\n     ↓\n Stop → kembali admin",
            {"admin": True},
            "Audit",
            "Sesuai kebutuhan",
            "Banner kuning + sesi review",
            "",
            None,
            80,
        ),
    ]
    for s in steps:
        r = add_step(ws, r, *s)
    footer_note(ws, r + 1, "*** SOP Penugasan — integrasi www/bun dan www/pengawas ***")


def sheet_akses_admin(wb: Workbook):
    ws = wb.create_sheet("SOP-06 Manajemen Akses")
    setup_columns(ws)
    r = write_header(ws, "FLOWCHART SOP MANAJEMEN AKSES & PERMISSION (ADMIN)")

    steps = [
        (
            "1",
            "Admin membuka Roles (/roles) dan Permissions (/permissions): "
            "pastikan peran standar tersedia.",
            "┌──────────┐\n│   Role   │\n└────┬─────┘\n     ↓",
            {"admin": True},
            "RBAC",
            "± 15 menit",
            "Role & permission siap",
            "",
            None,
            72,
        ),
        (
            "2",
            "Atur Route Permissions dan Menu Permissions sesuai kebijakan unit.",
            "┌──────────┐\n│  Route   │\n│   Menu   │\n└────┬─────┘\n     ↓",
            {"admin": True},
            "RBAC",
            "± 20 menit",
            "Akses modul terkontrol",
            "",
            None,
            80,
        ),
        (
            "3",
            "Tambah user baru di /users (nama, email, password, role). "
            "Opsional: batasi kegiatan di /kegiatan-role.",
            "┌──────────┐\n│   User   │\n└────┬─────┘\n     ↓",
            {"admin": True},
            "User mgmt",
            "± 10 menit",
            "Akun aktif",
            "",
            None,
            72,
        ),
        (
            "4",
            "Uji login sebagai user baru: pastikan menu sidebar sesuai permission.",
            "   ◇ Sesuai?\n  ╱        ╲\nTidak      Ya\n ↓          ↓\nPerbaiki   Selesai",
            {"admin": True},
            "Uji coba",
            "± 5 menit",
            "Hak akses benar",
            "",
            DECISION_FILL,
            88,
        ),
    ]
    for s in steps:
        r = add_step(ws, r, *s)
    footer_note(ws, r + 1, "*** SOP Manajemen Akses — Admin only ***")


def sheet_index(wb: Workbook):
    ws = wb.create_sheet("Daftar Isi", 0)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 14

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "DAFTAR ISI — SOP PENGGUNAAN ARUMANIS & PANEL PENGAWASAN"
    sty(c, bold=True, size=14, fill=TITLE_FILL, center=True)
    ws.row_dimensions[1].height = 30

    meta = [
        ("Versi dokumen", "1.0"),
        ("Platform", "Arumanis v0.5.0"),
        ("Tanggal", "1 Juli 2026"),
        ("Repo frontend", "www/bun (Arumanis) · www/pengawas (Panel Pengawasan)"),
        ("Backend", "apiamis (Laravel)"),
        ("URL produksi", "https://arumanis.cianjurkab.go.id"),
    ]
    row = 3
    for k, v in meta:
        ws.cell(row=row, column=1, value=k)
        sty(ws.cell(row=row, column=1), bold=True)
        ws.merge_cells(f"B{row}:D{row}")
        ws.cell(row=row, column=2, value=v)
        sty(ws.cell(row=row, column=2))
        row += 1

    row += 1
    for col, h in enumerate(["No", "Lembar SOP", "Judul Flowchart", "Aplikasi"], 1):
        sty(ws.cell(row=row, column=col, value=h), bold=True, fill=HEADER_FILL, center=True)
    row += 1

    index = [
        ("1", "SOP-01 Login & SSO", "Akses & Autentikasi Arumanis / Panel Pengawasan", "Keduanya"),
        ("2", "SOP-02 Input Program", "Input Program Baru (Kegiatan → Pekerjaan → Output)", "www/bun"),
        ("3", "SOP-03 Kontrak", "Pengelolaan Kontrak & Penyedia", "www/bun"),
        ("4", "SOP-04 Pengawasan Lapangan", "Pemantauan & Pelaporan Lapangan", "www/pengawas"),
        ("5", "SOP-05 Penugasan Pengawas", "Penugasan Admin → Panel Pengawasan", "Keduanya"),
        ("6", "SOP-06 Manajemen Akses", "Role, Permission, User (Admin)", "www/bun"),
    ]
    for no, sheet, title, app in index:
        ws.cell(row=row, column=1, value=no)
        ws.cell(row=row, column=2, value=sheet)
        ws.cell(row=row, column=3, value=title)
        ws.cell(row=row, column=4, value=app)
        for col in range(1, 5):
            sty(ws.cell(row=row, column=col), center=(col in (1, 2, 4)))
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1)
    c.value = (
        "Keterangan kolom PIC: ● = penanggung jawab utama pada langkah tersebut. "
        "Kolom Flowchart menggunakan simbol proses (┌─┐), keputusan (◇), dan panah (↓). "
        "Regenerasi: python scripts/generate-sop-xlsx.py"
    )
    sty(c, size=9, wrap=True)


def main():
    wb = Workbook()
    # remove default sheet after creating index
    default = wb.active
    sheet_index(wb)
    sheet_login(wb)
    sheet_input_program(wb)
    sheet_kontrak(wb)
    sheet_pengawasan(wb)
    sheet_penugasan(wb)
    sheet_akses_admin(wb)
    if default.title == "Sheet":
        wb.remove(default)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUT)
        print(f"✓ SOP Excel tersimpan: {OUT}")
    except PermissionError:
        alt = OUT.with_stem(OUT.stem + "_baru")
        wb.save(alt)
        print(f"⚠ File terkunci. Disimpan ke: {alt}")


if __name__ == "__main__":
    main()