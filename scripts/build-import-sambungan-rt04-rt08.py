"""Build template import Sambungan Rumah from DAFTAR PENERIMA RT04 + RT08 PDFs.

Foto hanya level 0% (sambungan rumah 0%).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

TEMPLATE = Path(r"C:\Users\asusg\Downloads\template_import_sambungan_rumah (1).xlsx")
OUT = Path(r"C:\Users\asusg\Downloads\template_import_sambungan_rumah_RT04_RT08_ciloto.xlsx")
OUT_DOCS = Path(r"C:\laragon\www\bun\docs\template_import_sambungan_rumah_RT04_RT08_ciloto.xlsx")

# Parsed from DAFTAR PERIMA RT08.pdf / RT04.pdf (OCR structured sections).
# NIK: 16 digit (digit nomor baris yang menempel di OCR sudah dipotong).
# Foto: hanya nama_file_foto_0 — foto adalah sambungan rumah 0%.

RT08: list[tuple] = [
    (1, "Husein", "3275091703840034", "Kp. Parabon Pojok RT. 008/003", 3, -6.717647, 107.004731),
    (2, "Dodi Rahman", "3203281006040002", "Kp. Parabon Pojok RT. 008/003", 2, -6.717666, 107.004747),
    (3, "Anah", "3203284508500004", "Kp. Parabon Pojok RT. 008/003", 1, -6.717656, 107.004735),
    (4, "Ardi Sumardi", "3203281805630001", "Kp. Parabon Pojok RT. 008/003", 2, -6.717076, 107.004767),
    (5, "Hendi", "3203281208720007", "Kp. Parabon Pojok RT. 008/003", 4, -6.717088, 107.004772),
    (6, "Budi Supriadi", "3201252708780001", "Kp. Parabon Pojok RT. 008/003", 5, -6.71689, 107.005785),
    (7, "Deden", "3203280805970009", "Kp. Parabon Pojok RT. 008/003", 3, -6.717022, 107.004856),
    (8, "Aat", "3203280301890003", "Kp. Parabon Pojok RT. 008/003", 3, -6.716949, 107.004622),
    (9, "Jujun Junaedi", "3203281111770008", "Kp. Parabon Pojok RT. 008/003", 4, -6.717285, 107.00583),
    (10, "Didih S.", "3202391907810004", "Kp. Parabon Pojok RT. 008/003", 5, -6.717327, 107.004815),
    (11, "M. Alwin Niawan", "3203282203820008", "Kp. Parabon Pojok RT. 008/003", 4, -6.717072, 107.004758),
    (12, "Sukandi", "3203280909770007", "Kp. Parabon Pojok RT. 008/003", 5, -6.716638, 107.007024),
    (13, "Acep Mulyadi", "3203282206850003", "Kp. Parabon Pojok RT. 008/003", 4, -6.717178, 107.005827),
    (14, "Imas Herawati", "3203284207630007", "Kp. Parabon Pojok RT. 008/003", 2, -6.716907, 107.005103),
    (15, "Suryana", "3203280301920009", "Kp. Parabon Pojok RT. 008/003", 5, -6.717128, 107.006372),
    (16, "Ade Wahyu", "3203280408700003", "Kp. Parabon Pojok RT. 008/003", 3, -6.717128, 107.005292),
    (17, "Rahmatulloh", "3203281101830003", "Kp. Parabon Pojok RT. 008/003", 5, -6.71693, 107.004747),
    (18, "Badrudin", "3203281207730005", "Kp. Parabon Pojok RT. 008/003", 5, -6.716373, 107.006295),
    (19, "Saepul Rohmat", "3203282709770003", "Kp. Parabon Pojok RT. 008/003", 4, -6.717382, 107.005027),
    (20, "Yarti Herawati", "3203285210710002", "Kp. Parabon Pojok RT. 008/003", 4, -6.716743, 107.006),
    (21, "Djemain", "3203282109700004", "Kp. Parabon Pojok RT. 008/003", 3, -6.717354, 107.004431),
    (22, "Rahmat", "3203280606600024", "Kp. Parabon Pojok RT. 008/003", 4, -6.717288, 107.00526),
    (23, "Hm. Faqih Arrozi", "3203281511700002", "Kp. Parabon Pojok RT. 008/003", 3, -6.716464, 107.006288),
    (24, "Iman", "3202010609910007", "Kp. Parabon Pojok RT. 008/003", 4, -6.717036, 107.004829),
    (25, "Adam", "3203281604950004", "Kp. Parabon Pojok RT. 008/003", 3, -6.716944, 107.004653),
    (26, "Imam Rustandi", "3203282105800007", "Kp. Parabon Pojok RT. 008/003", 5, -6.717647, 107.004756),
    (27, "Evi Nuryati", "3203284704900006", "Kp. Parabon Pojok RT. 008/003", 4, -6.717156, 107.005863),
    (28, "Yuda Herlambang", "3203281305800003", "Kp. Parabon Pojok RT. 008/003", 3, -6.716731, 107.00605),
    (29, "Dadang", "3203281407000052", "Kp. Parabon Pojok RT. 008/003", 3, -6.717402, 107.004945),
]

RT04: list[tuple] = [
    (1, "Hadiyono", "3203280702740001", "Kp. Parabon Pojok RT. 004/003", 4, -6.717565, 107.003706),
    (2, "Saepul Ajis", "3203280704820012", "Kp. Parabon Pojok RT. 004/003", 4, -6.717168, 107.003344),
    (3, "Pujino", "3203281911640003", "Kp. Parabon Pojok RT. 004/003", 3, -6.717599, 107.004029),
    (4, "Ahmad Fadlullah", "3203281303730002", "Kp. Parabon Pojok RT. 004/003", 4, -6.717538, 107.002973),
    (5, "Ladi Sunandar", "3203280502890008", "Kp. Parabon Pojok RT. 004/003", 3, -6.717422, 107.004056),
    (6, "Teguh Santosa", "3276071408880001", "Kp. Parabon Pojok RT. 004/003", 4, -6.71715, 107.00413),
    (7, "Mahpudin", "3203281011690003", "Kp. Parabon Pojok RT. 004/003", 5, -6.717322, 107.00281),
    (8, "Hermansah", "3203281406790017", "Kp. Parabon Pojok RT. 004/003", 4, -6.717306, 107.002662),
    (9, "Asep Dirman", "3203281507840008", "Kp. Parabon Pojok RT. 004/003", 4, -6.717705, 107.003033),
    (10, "Jamili", "3203280101708511", "Kp. Parabon Pojok RT. 004/003", 3, -6.717334, 107.002752),
    (11, "Rohman", "3203280304830002", "Kp. Parabon Pojok RT. 004/003", 5, -6.717429, 107.003396),
    (12, "Yani Agus Sujangi", "3203100107781354", "Kp. Parabon Pojok RT. 004/003", 3, -6.717416, 107.003413),
    (13, "Sutisna", "3203102504733000", "Kp. Parabon Pojok RT. 004/003", 4, -6.717479, 107.003145),
    (14, "Dudum Dumyati", "3203281205870008", "Kp. Parabon Pojok RT. 004/003", 4, -6.717423, 107.004055),
    (15, "Budi Haryadi", "3203282209780006", "Kp. Parabon Pojok RT. 004/003", 4, -6.717516, 107.004066),
    (16, "Ade Sopian", "3203280212850003", "Kp. Parabon Pojok RT. 004/003", 4, -6.71734, 107.003826),
    (17, "Bubun Setiawan", "3203280808800001", "Kp. Parabon Pojok RT. 004/003", 4, -6.717713, 107.003612),
    (18, "Cecep Abdul Patah", "3203281503760004", "Kp. Parabon Pojok RT. 004/003", 6, -6.717499, 107.003554),
    (19, "Dedi Suhendi", "3203280107600038", "Kp. Parabon Pojok RT. 004/003", 3, -6.717561, 107.003156),
    (20, "Riki", "3203281508840016", "Kp. Parabon Pojok RT. 004/003", 3, -6.717533, 107.002987),
    (21, "Bambang", "3203281509750001", "Kp. Parabon Pojok RT. 004/003", 3, -6.717603, 107.004055),
    (22, "Sukirman Ardiansah", "3203281206890014", "Kp. Parabon Pojok RT. 004/003", 4, -6.717433, 107.004046),
    (23, "Ujang Hudori", "3203282411880005", "Kp. Parabon Pojok RT. 004/003", 5, -6.717536, 107.003436),
    (24, "Rostika", "3203284101740060", "Kp. Parabon Pojok RT. 004/003", 3, -6.717372, 107.002806),
    (25, "Yanto", "3203281110800012", "Kp. Parabon Pojok RT. 004/003", 4, -6.717072, 107.004267),
    (26, "Rosmiati", "3203104102800030", "Kp. Parabon Pojok RT. 004/003", 4, -6.717647, 107.004756),
]


def title_case_name(name: str) -> str:
    parts = []
    for p in name.split():
        if p.upper() in {"S.", "M.", "HM."}:
            parts.append(p.upper() if p.endswith(".") else p.title())
        else:
            parts.append(p.title())
    return " ".join(parts)


def main() -> None:
    all_rows: list[tuple] = []
    for no, nama, nik, alamat, jiwa, lat, lon in RT08:
        assert len(nik) == 16, (no, nik)
        all_rows.append((nama, nik, alamat, jiwa, lat, lon, "RT08", no))
    for no, nama, nik, alamat, jiwa, lat, lon in RT04:
        assert len(nik) == 16, (no, nik)
        all_rows.append((nama, nik, alamat, jiwa, lat, lon, "RT04", no))

    wb = openpyxl.load_workbook(TEMPLATE)
    ws_info = wb["Info"]
    ws_data = wb["Data"]
    ws_petunjuk = wb["Petunjuk"]

    if ws_data.max_row > 1:
        ws_data.delete_rows(2, ws_data.max_row - 1)

    total = len(all_rows)
    ws_info["B2"] = "Sambungan Rumah"
    ws_info["C2"] = "Unit"
    ws_info["D2"] = float(total)
    ws_info["E2"] = "unit"
    ws_info["F2"] = total

    ws_petunjuk["A1"] = "TEMPLATE IMPORT UNIT — Sambungan Rumah (RT04 + RT08 Desa Ciloto)"
    ws_petunjuk["A8"] = (
        f"Target baris penerima: {total} ({total}.00 Unit). "
        "Sumber: DAFTAR PENERIMA RT08 (29) + RT04 (26). "
        "Foto hanya 0% (sambungan rumah 0%) di kolom nama_file_foto_0."
    )
    ws_petunjuk["A15"] = (
        "Catatan NIK: beberapa NIK dari PDF digabung dengan nomor urut OCR; "
        "sudah dinormalisasi ke 16 digit. Verifikasi ulang NIK di lapangan bila perlu."
    )
    ws_petunjuk["A16"] = (
        "Urutan baris: 1–29 = RT.008/003, 30–55 = RT.004/003. "
        "ZIP foto contoh: 001_0.jpg … 055_0.jpg."
    )

    for i, (nama, nik, alamat, jiwa, lat, lon, _rt, _src_no) in enumerate(all_rows, start=1):
        ws_data.append(
            [
                i,
                title_case_name(nama),
                nik,
                alamat,
                jiwa,
                lat,
                lon,
                f"{i:03d}_0.jpg",  # hanya foto 0%
                None,
                None,
                None,
                None,
            ]
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"Total penerima: {total} (RT08={len(RT08)}, RT04={len(RT04)})")
    print("Kolom foto: hanya nama_file_foto_0 (sambungan rumah 0%)")

    try:
        OUT_DOCS.parent.mkdir(parents=True, exist_ok=True)
        wb.save(OUT_DOCS)
        print(f"Also: {OUT_DOCS}")
    except OSError as exc:
        print(f"Skip docs copy: {exc}")


if __name__ == "__main__":
    main()
