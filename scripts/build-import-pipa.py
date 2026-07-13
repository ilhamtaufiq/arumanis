"""Build import package for komponen Pipa (komunal) from WhatsApp photos.

- Template: template_import_pipa.xlsx (komunal, satuan Meter)
- Foto: WhatsApp zip — tiap titik GPS jadi unit_index (segmen dokumentasi)
- Level foto: 0% per unit (satu foto utama); jika beberapa foto di titik yang sama
  diurutkan waktu → diisi 0/25/50/75/100
"""

from __future__ import annotations

import io
import math
import re
import zipfile
from pathlib import Path

import easyocr
import numpy as np
import openpyxl
from PIL import Image

TEMPLATE = Path(r"C:\Users\asusg\Downloads\template_import_pipa.xlsx")
WA_ZIP = Path(r"C:\Users\asusg\Downloads\WhatsApp Unknown 2026-07-09 at 11.29.26.zip")
OUT_XLSX = Path(r"C:\Users\asusg\Downloads\template_import_pipa_ciloto.xlsx")
OUT_ZIP = Path(r"C:\Users\asusg\Downloads\foto_pipa_ciloto.zip")
OUT_DIR = Path(r"C:\Users\asusg\Downloads\foto_pipa_ciloto")
OUT_XLSX_DOCS = Path(r"C:\laragon\www\bun\docs\template_import_pipa_ciloto.xlsx")

LEVELS = ["0", "25", "50", "75", "100"]
CLUSTER_M = 35.0  # radius klaster titik dokumentasi
LAT_LONG_RE = re.compile(
    r"Lat\s*[-–]?\s*(\d+[.,]\d+).{0,40}Long\s*[-–]?\s*(\d+[.,]\d+)",
    re.I | re.S,
)


def wa_sort_key(name: str) -> tuple:
    m = re.search(r"at (\d{2})\.(\d{2})\.(\d{2})(?: \((\d+)\))?\.jpe?g", name, re.I)
    if not m:
        return (name,)
    return (int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4) or 0))


def haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    r = 6371000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlmb / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def parse_lat_lon(text: str) -> tuple[float, float] | None:
    text = text.replace(",", ".")
    m = LAT_LONG_RE.search(text)
    if not m:
        return None
    lat = float(m.group(1))
    lon = float(m.group(2))
    if lat > 0:
        lat = -lat
    if not (-7.5 < lat < -6.0 and 106.5 < lon < 107.5):
        return None
    return lat, lon


def ocr_gps(reader: easyocr.Reader, img: Image.Image) -> tuple[float, float] | None:
    w, h = img.size
    crops = [
        img.crop((int(w * 0.15), int(h * 0.70), w, h)),
        img.crop((int(w * 0.25), int(h * 0.78), w, h)),
    ]
    texts: list[str] = []
    for crop in crops:
        try:
            parts = reader.readtext(np.array(crop), detail=0, paragraph=False)
            texts.append(" ".join(str(p) for p in parts))
        except Exception:
            continue
    return parse_lat_lon(" | ".join(texts))


def save_jpeg(img: Image.Image, path: Path, quality: int = 88) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if img.mode != "RGB":
        img = img.convert("RGB")
    img.save(path, format="JPEG", quality=quality, optimize=True)


def cluster_photos(photos: list[dict], radius_m: float) -> list[list[dict]]:
    """Greedy spatial clusters, order by first photo time."""
    clusters: list[list[dict]] = []
    for photo in photos:
        if not photo["gps"]:
            clusters.append([photo])
            continue
        lat, lon = photo["gps"]
        placed = False
        for cluster in clusters:
            # centroid of cluster
            gps_list = [p["gps"] for p in cluster if p["gps"]]
            if not gps_list:
                continue
            clat = sum(g[0] for g in gps_list) / len(gps_list)
            clon = sum(g[1] for g in gps_list) / len(gps_list)
            if haversine_m(lat, lon, clat, clon) <= radius_m:
                cluster.append(photo)
                placed = True
                break
        if not placed:
            clusters.append([photo])
    return clusters


def main() -> None:
    print("Load EasyOCR…")
    reader = easyocr.Reader(["en"], gpu=False, verbose=False)

    print(f"Read photos from {WA_ZIP.name}…")
    photos: list[dict] = []
    with zipfile.ZipFile(WA_ZIP) as zf:
        names = sorted(
            [n for n in zf.namelist() if n.lower().endswith((".jpg", ".jpeg", ".png"))],
            key=wa_sort_key,
        )
        for i, name in enumerate(names, start=1):
            raw = zf.read(name)
            img = Image.open(io.BytesIO(raw)).convert("RGB")
            gps = ocr_gps(reader, img)
            photos.append({"name": name, "img": img, "gps": gps, "order": i})
            gps_s = f"{gps[0]:.6f},{gps[1]:.6f}" if gps else "NO_GPS"
            print(f"  [{i:02d}/{len(names)}] {Path(name).name[:50]} {gps_s}")

    with_gps = sum(1 for p in photos if p["gps"])
    print(f"Total foto: {len(photos)}, GPS: {with_gps}")

    # Satu foto = satu unit dokumentasi (trase pipa), level 0% (+100% jika ada pasangan dekat)
    # Agar semua foto terpakai, tidak digabung max-5 per klaster.
    print("Mode: 1 foto → 1 unit (nama_file_foto_0)")
    units: list[dict] = []
    for ci, photo in enumerate(photos, start=1):
        lat = lon = None
        if photo["gps"]:
            lat, lon = photo["gps"]
        fname0 = f"{ci:03d}_0.jpg"
        units.append(
            {
                "unit_index": ci,
                "label": f"Segmen Pipa {ci}",
                "lat": lat,
                "lon": lon,
                "files": {"0": fname0},
                "images": {"0": photo["img"]},
            }
        )
        gps_s = f"{lat:.6f},{lon:.6f}" if lat is not None else "NO_GPS"
        print(f"  Unit {ci:02d}: {fname0}  {gps_s}")

    # --- Write images + ZIP ---
    if OUT_DIR.exists():
        for old in OUT_DIR.glob("*.jpg"):
            old.unlink()
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    zip_names: list[str] = []
    for unit in units:
        for level, fname in unit["files"].items():
            save_jpeg(unit["images"][level], OUT_DIR / fname)
            zip_names.append(fname)

    if OUT_ZIP.exists():
        OUT_ZIP.unlink()
    with zipfile.ZipFile(OUT_ZIP, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name in zip_names:
            zf.write(OUT_DIR / name, arcname=name)

    # --- Excel ---
    wb = openpyxl.load_workbook(TEMPLATE)
    ws_info = wb["Info"]
    ws_data = wb["Data"]
    ws_petunjuk = wb["Petunjuk"]

    if ws_data.max_row > 1:
        ws_data.delete_rows(2, ws_data.max_row - 1)

    n = len(units)
    # keep komponen_id, komponen, satuan, volume from template
    ws_info["E2"] = "komunal"
    ws_info["F2"] = n  # target_baris = jumlah unit dokumentasi

    ws_petunjuk["A1"] = "TEMPLATE IMPORT KOMUNAL — Pipa (Desa Ciloto)"
    ws_petunjuk["A9"] = (
        f"Target baris/unit: {n} (volume pipa tetap {ws_info['D2']} {ws_info['C2']}). "
        f"Unit = titik dokumentasi di sepanjang trase (klaster GPS ≤ {int(CLUSTER_M)} m)."
    )
    ws_petunjuk["A15"] = (
        f"Sumber: {WA_ZIP.name} ({len(photos)} foto). "
        "Setiap foto = 1 unit_index (segmen trase), kolom nama_file_foto_0."
    )
    ws_petunjuk["A16"] = (
        "Volume Info tetap total meter pipa; target_baris = jumlah titik dokumentasi foto."
    )

    for i, unit in enumerate(units, start=1):
        files = unit["files"]
        ws_data.append(
            [
                i,
                unit["unit_index"],
                unit["label"],
                unit["lat"],
                unit["lon"],
                files.get("0"),
                files.get("25"),
                files.get("50"),
                files.get("75"),
                files.get("100"),
            ]
        )

    wb.save(OUT_XLSX)
    try:
        OUT_XLSX_DOCS.parent.mkdir(parents=True, exist_ok=True)
        wb.save(OUT_XLSX_DOCS)
    except OSError:
        pass

    size_mb = OUT_ZIP.stat().st_size / (1024 * 1024)
    print()
    print(f"Excel: {OUT_XLSX}")
    print(f"ZIP:   {OUT_ZIP} ({size_mb:.2f} MB), files={len(zip_names)}")
    print(f"Dir:   {OUT_DIR}")
    print(f"Units: {n}, photos used: {len(zip_names)}/{len(photos)}")


if __name__ == "__main__":
    main()
