"""Rebuild import Sambungan Rumah: penerima + foto 0% (PDF) + foto 100% (WhatsApp zip).

Matching foto 100% ke penerima lewat koordinat watermark GPS Map Camera (EasyOCR).
"""

from __future__ import annotations

import io
import math
import re
import zipfile
from pathlib import Path

import easyocr
import fitz
import numpy as np
import openpyxl
from PIL import Image

# --- Paths ---
TEMPLATE = Path(r"C:\Users\asusg\Downloads\template_import_sambungan_rumah (1).xlsx")
RT08_PDF = Path(r"C:\Users\asusg\Downloads\DAFTAR PERIMA RT08.pdf")
RT04_PDF = Path(r"C:\Users\asusg\Downloads\DAFTAR PERIMA RT04.pdf")
WA_ZIP = Path(r"C:\Users\asusg\Downloads\WhatsApp Unknown 2026-07-09 at 10.46.21.zip")

OUT_XLSX = Path(r"C:\Users\asusg\Downloads\template_import_sambungan_rumah_RT04_RT08_ciloto.xlsx")
OUT_ZIP = Path(r"C:\Users\asusg\Downloads\foto_sambungan_rumah_0_100_RT04_RT08_ciloto.zip")
OUT_DIR = Path(r"C:\Users\asusg\Downloads\foto_sambungan_rumah_0_100_RT04_RT08_ciloto")
OUT_XLSX_DOCS = Path(r"C:\laragon\www\bun\docs\template_import_sambungan_rumah_RT04_RT08_ciloto.xlsx")

# --- Penerima (dari OCR PDF sebelumnya) ---
RT08: list[tuple] = [
    (1, "Husein", "3275091703840034", "Kp. Parabon Pojok RT. 008/003", 3, -6.717647, 107.004731),
    (2, "Dodi Rahman", "3203281006040002", "Kp. Parabon Pojok RT. 008/003", 2, -6.717666, 107.004747),
    (3, "Anah", "3203284508500004", "Kp. Parabon Pojok RT. 008/003", 1, -6.717656, 107.004735),
    (4, "Ardi Sumardi", "3203281805630001", "Kp. Parabon Pojok RT. 008/003", 2, -6.717076, 107.004767),
    (5, "Hendi", "3203281208720007", "Kp. Parabon Pojok RT. 008/003", 4, -6.717088, 107.004772),
    (6, "Budi Supriadi", "3201252708780001", "Kp. Parabon Pojok RT. 008/003", 5, -6.71689, 107.005785),
    (7, "Deden", "3203280805970009", "Kp. Parabon Pojok RT. 008/003", 3, -6.717022, 107.004856),
    (8, "Aat", "3203280301890003", "Kp. Parabon Pojok RT. 008/003", 3, -6.716949, 107.004622),
    (9, "Jujun Junaedi", "3203281111770008", "Kp. Parabon Pojok RT. 008/003", 4, -6.717285, 107.00583),
    (10, "Didih S.", "3202391907810004", "Kp. Parabon Pojok RT. 008/003", 5, -6.717327, 107.004815),
    (11, "M. Alwin Niawan", "3203282203820008", "Kp. Parabon Pojok RT. 008/003", 4, -6.717072, 107.004758),
    (12, "Sukandi", "3203280909770007", "Kp. Parabon Pojok RT. 008/003", 5, -6.716638, 107.007024),
    (13, "Acep Mulyadi", "3203282206850003", "Kp. Parabon Pojok RT. 008/003", 4, -6.717178, 107.005827),
    (14, "Imas Herawati", "3203284207630007", "Kp. Parabon Pojok RT. 008/003", 2, -6.716907, 107.005103),
    (15, "Suryana", "3203280301920009", "Kp. Parabon Pojok RT. 008/003", 5, -6.717128, 107.006372),
    (16, "Ade Wahyu", "3203280408700003", "Kp. Parabon Pojok RT. 008/003", 3, -6.717128, 107.005292),
    (17, "Rahmatulloh", "3203281101830003", "Kp. Parabon Pojok RT. 008/003", 5, -6.71693, 107.004747),
    (18, "Badrudin", "3203281207730005", "Kp. Parabon Pojok RT. 008/003", 5, -6.716373, 107.006295),
    (19, "Saepul Rohmat", "3203282709770003", "Kp. Parabon Pojok RT. 008/003", 4, -6.717382, 107.005027),
    (20, "Yarti Herawati", "3203285210710002", "Kp. Parabon Pojok RT. 008/003", 4, -6.716743, 107.006),
    (21, "Djemain", "3203282109700004", "Kp. Parabon Pojok RT. 008/003", 3, -6.717354, 107.004431),
    (22, "Rahmat", "3203280606600024", "Kp. Parabon Pojok RT. 008/003", 4, -6.717288, 107.00526),
    (23, "Hm. Faqih Arrozi", "3203281511700002", "Kp. Parabon Pojok RT. 008/003", 3, -6.716464, 107.006288),
    (24, "Iman", "3202010609910007", "Kp. Parabon Pojok RT. 008/003", 4, -6.717036, 107.004829),
    (25, "Adam", "3203281604950004", "Kp. Parabon Pojok RT. 008/003", 3, -6.716944, 107.004653),
    (26, "Imam Rustandi", "3203282105800007", "Kp. Parabon Pojok RT. 008/003", 5, -6.717647, 107.004756),
    (27, "Evi Nuryati", "3203284704900006", "Kp. Parabon Pojok RT. 008/003", 4, -6.717156, 107.005863),
    (28, "Yuda Herlambang", "3203281305800003", "Kp. Parabon Pojok RT. 008/003", 3, -6.716731, 107.00605),
    (29, "Dadang", "3203281407000052", "Kp. Parabon Pojok RT. 008/003", 3, -6.717402, 107.004945),
]

RT04: list[tuple] = [
    (1, "Hadiyono", "3203280702740001", "Kp. Parabon Pojok RT. 004/003", 4, -6.717565, 107.003706),
    (2, "Saepul Ajis", "3203280704820012", "Kp. Parabon Pojok RT. 004/003", 4, -6.717168, 107.003344),
    (3, "Pujino", "3203281911640003", "Kp. Parabon Pojok RT. 004/003", 3, -6.717599, 107.004029),
    (4, "Ahmad Fadlullah", "3203281303730002", "Kp. Parabon Pojok RT. 004/003", 4, -6.717538, 107.002973),
    (5, "Ladi Sunandar", "3203280502890008", "Kp. Parabon Pojok RT. 004/003", 3, -6.717422, 107.004056),
    (6, "Teguh Santosa", "3276071408880001", "Kp. Parabon Pojok RT. 004/003", 4, -6.71715, 107.00413),
    (7, "Mahpudin", "3203281011690003", "Kp. Parabon Pojok RT. 004/003", 5, -6.717322, 107.00281),
    (8, "Hermansah", "3203281406790017", "Kp. Parabon Pojok RT. 004/003", 4, -6.717306, 107.002662),
    (9, "Asep Dirman", "3203281507840008", "Kp. Parabon Pojok RT. 004/003", 4, -6.717705, 107.003033),
    (10, "Jamili", "3203280101708511", "Kp. Parabon Pojok RT. 004/003", 3, -6.717334, 107.002752),
    (11, "Rohman", "3203280304830002", "Kp. Parabon Pojok RT. 004/003", 5, -6.717429, 107.003396),
    (12, "Yani Agus Sujangi", "3203100107781354", "Kp. Parabon Pojok RT. 004/003", 3, -6.717416, 107.003413),
    (13, "Sutisna", "3203102504733000", "Kp. Parabon Pojok RT. 004/003", 4, -6.717479, 107.003145),
    (14, "Dudum Dumyati", "3203281205870008", "Kp. Parabon Pojok RT. 004/003", 4, -6.717423, 107.004055),
    (15, "Budi Haryadi", "3203282209780006", "Kp. Parabon Pojok RT. 004/003", 4, -6.717516, 107.004066),
    (16, "Ade Sopian", "3203280212850003", "Kp. Parabon Pojok RT. 004/003", 4, -6.71734, 107.003826),
    (17, "Bubun Setiawan", "3203280808800001", "Kp. Parabon Pojok RT. 004/003", 4, -6.717713, 107.003612),
    (18, "Cecep Abdul Patah", "3203281503760004", "Kp. Parabon Pojok RT. 004/003", 6, -6.717499, 107.003554),
    (19, "Dedi Suhendi", "3203280107600038", "Kp. Parabon Pojok RT. 004/003", 3, -6.717561, 107.003156),
    (20, "Riki", "3203281508840016", "Kp. Parabon Pojok RT. 004/003", 3, -6.717533, 107.002987),
    (21, "Bambang", "3203281509750001", "Kp. Parabon Pojok RT. 004/003", 3, -6.717603, 107.004055),
    (22, "Sukirman Ardiansah", "3203281206890014", "Kp. Parabon Pojok RT. 004/003", 4, -6.717433, 107.004046),
    (23, "Ujang Hudori", "3203282411880005", "Kp. Parabon Pojok RT. 004/003", 5, -6.717536, 107.003436),
    (24, "Rostika", "3203284101740060", "Kp. Parabon Pojok RT. 004/003", 3, -6.717372, 107.002806),
    (25, "Yanto", "3203281110800012", "Kp. Parabon Pojok RT. 004/003", 4, -6.717072, 107.004267),
    (26, "Rosmiati", "3203104102800030", "Kp. Parabon Pojok RT. 004/003", 4, -6.717647, 107.004756),
]

LAT_LONG_RE = re.compile(
    r"Lat\s*[-–]?\s*(\d+[.,]\d+)\s*°?\s*Long\s*[-–]?\s*(\d+[.,]\d+)",
    re.IGNORECASE,
)
# OCR often drops the minus: "Lat 6.7173970 Long 107.0048840"
LAT_LONG_RE2 = re.compile(
    r"Lat\s*[-–]?\s*(\d+[.,]\d+).{0,20}Long\s*[-–]?\s*(\d+[.,]\d+)",
    re.IGNORECASE | re.DOTALL,
)


def title_case_name(name: str) -> str:
    parts = []
    for p in name.split():
        if p.upper() in {"S.", "M.", "HM."}:
            parts.append(p.upper())
        else:
            parts.append(p.title())
    return " ".join(parts)


def haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    r = 6371000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlmb / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def extract_pdf_photos(pdf_path: Path) -> list[Image.Image]:
    doc = fitz.open(pdf_path)
    photos: list[Image.Image] = []
    for i in range(doc.page_count):
        page = doc[i]
        blocks = [b for b in page.get_text("dict")["blocks"] if b.get("type") == 1]
        blocks.sort(key=lambda b: (round(b["bbox"][1], 1), round(b["bbox"][0], 1)))
        for block in blocks:
            raw = block.get("image")
            if raw:
                img = Image.open(io.BytesIO(raw))
            else:
                rect = fitz.Rect(block["bbox"])
                pix = page.get_pixmap(clip=rect, matrix=fitz.Matrix(2, 2), alpha=False)
                img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
            if img.mode != "RGB":
                img = img.convert("RGB")
            photos.append(img)
        print(f"  {pdf_path.name} p{i + 1}: {len(blocks)} foto")
    doc.close()
    return photos


def wa_sort_key(name: str) -> tuple:
    m = re.search(r"at (\d{2})\.(\d{2})\.(\d{2})(?: \((\d+)\))?\.jpe?g", name, re.I)
    if not m:
        return (name,)
    return (int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4) or 0))


def parse_lat_lon_from_text(text: str) -> tuple[float, float] | None:
    text = text.replace(",", ".")
    m = LAT_LONG_RE.search(text) or LAT_LONG_RE2.search(text)
    if not m:
        return None
    lat = float(m.group(1))
    lon = float(m.group(2))
    # Ciloto always south of equator
    if lat > 0:
        lat = -lat
    if not (-7.5 < lat < -6.0 and 106.5 < lon < 107.5):
        return None
    return lat, lon


def ocr_photo_gps(reader: easyocr.Reader, img: Image.Image) -> tuple[float, float] | None:
    w, h = img.size
    # Watermark is bottom-right / bottom band
    crops = [
        img.crop((int(w * 0.20), int(h * 0.72), w, h)),
        img.crop((int(w * 0.30), int(h * 0.78), w, h)),
    ]
    texts: list[str] = []
    for crop in crops:
        arr = np.array(crop)
        try:
            parts = reader.readtext(arr, detail=0, paragraph=False)
        except Exception:
            continue
        texts.append(" ".join(str(p) for p in parts))
    joined = " | ".join(texts)
    return parse_lat_lon_from_text(joined)


def save_jpeg(img: Image.Image, path: Path, quality: int = 90) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if img.mode != "RGB":
        img = img.convert("RGB")
    img.save(path, format="JPEG", quality=quality, optimize=True)


def build_rows() -> list[dict]:
    rows: list[dict] = []
    for no, nama, nik, alamat, jiwa, lat, lon in RT08:
        rows.append(
            {
                "nama": title_case_name(nama),
                "nik": nik,
                "alamat": alamat,
                "jumlah_jiwa": jiwa,
                "lat": lat,
                "lon": lon,
                "rt": "RT08",
            }
        )
    for no, nama, nik, alamat, jiwa, lat, lon in RT04:
        rows.append(
            {
                "nama": title_case_name(nama),
                "nik": nik,
                "alamat": alamat,
                "jumlah_jiwa": jiwa,
                "lat": lat,
                "lon": lon,
                "rt": "RT04",
            }
        )
    return rows


def main() -> None:
    rows = build_rows()
    n = len(rows)
    print(f"Penerima: {n} (RT08={len(RT08)}, RT04={len(RT04)})")

    # --- Foto 0% dari PDF ---
    print("Extract foto 0% PDF…")
    foto0 = extract_pdf_photos(RT08_PDF) + extract_pdf_photos(RT04_PDF)
    if len(foto0) != n:
        raise SystemExit(f"Foto 0% count {len(foto0)} != penerima {n}")

    # --- Foto 100% WhatsApp + GPS OCR ---
    print("Load EasyOCR…")
    reader = easyocr.Reader(["en"], gpu=False, verbose=False)

    print(f"OCR GPS foto 100% dari {WA_ZIP.name}…")
    wa_photos: list[dict] = []
    with zipfile.ZipFile(WA_ZIP) as zf:
        names = sorted(
            [x for x in zf.namelist() if x.lower().endswith((".jpg", ".jpeg", ".png"))],
            key=wa_sort_key,
        )
        for i, name in enumerate(names, start=1):
            raw = zf.read(name)
            img = Image.open(io.BytesIO(raw))
            if img.mode != "RGB":
                img = img.convert("RGB")
            gps = ocr_photo_gps(reader, img)
            wa_photos.append({"name": name, "img": img, "gps": gps, "size": img.size})
            status = f"{gps[0]:.6f},{gps[1]:.6f}" if gps else "NO_GPS"
            print(f"  [{i:02d}/{len(names)}] {Path(name).name[:48]} {img.size} {status}")

    with_gps = [p for p in wa_photos if p["gps"]]
    print(f"WA foto: {len(wa_photos)}, dengan GPS OCR: {len(with_gps)}")

    # Prefer completed SR box photos (typically 1600x1200) over trench/high-res when both match
    def photo_priority(p: dict) -> tuple:
        w, h = p["size"]
        # lower is better: prefer landscape 1600x1200-ish, penalize huge dig photos slightly less than no match
        is_std = 1 if (w, h) in {(1600, 1200), (1200, 1600)} else 0
        is_hires = 1 if max(w, h) >= 3000 else 0
        return (0 if is_std else 1, 0 if not is_hires else 1)

    # Greedy assign: for each recipient, best unused photo within 80m
    MAX_DIST_M = 80.0
    assigned_100: dict[int, dict] = {}
    used_photo: set[int] = set()

    # Build candidate pairs
    pairs: list[tuple[float, int, int]] = []  # dist, row_idx, photo_idx
    for ri, row in enumerate(rows):
        for pi, photo in enumerate(wa_photos):
            if not photo["gps"]:
                continue
            d = haversine_m(row["lat"], row["lon"], photo["gps"][0], photo["gps"][1])
            if d <= MAX_DIST_M:
                # adjust score by photo priority
                prio = photo_priority(photo)
                score = d + prio[0] * 5 + prio[1] * 2
                pairs.append((score, ri, pi))
    pairs.sort(key=lambda x: x[0])

    for score, ri, pi in pairs:
        if ri in assigned_100 or pi in used_photo:
            continue
        assigned_100[ri] = wa_photos[pi]
        used_photo.add(pi)
        d = haversine_m(
            rows[ri]["lat"],
            rows[ri]["lon"],
            wa_photos[pi]["gps"][0],
            wa_photos[pi]["gps"][1],
        )
        print(f"  match #{ri + 1:02d} {rows[ri]['nama'][:20]:20s} <- {Path(wa_photos[pi]['name']).name[:40]} ({d:.1f}m)")

    unmatched = [i for i in range(n) if i not in assigned_100]
    print(f"Matched 100%: {len(assigned_100)}/{n}, unmatched: {len(unmatched)}")
    if unmatched:
        print("  unmatched:", ", ".join(f"{i + 1}:{rows[i]['nama']}" for i in unmatched))

    # Remaining photos by order → fill unmatched (fallback sequential)
    remaining = [i for i in range(len(wa_photos)) if i not in used_photo]
    # Prefer standard size for fallback
    remaining.sort(key=lambda i: photo_priority(wa_photos[i]))
    for ri in unmatched:
        if not remaining:
            break
        pi = remaining.pop(0)
        assigned_100[ri] = wa_photos[pi]
        used_photo.add(pi)
        print(f"  fallback #{ri + 1:02d} {rows[ri]['nama'][:20]:20s} <- {Path(wa_photos[pi]['name']).name[:40]}")

    # --- Write folder + ZIP ---
    if OUT_DIR.exists():
        for old in OUT_DIR.glob("*.jpg"):
            old.unlink()
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    zip_names: list[str] = []
    for i in range(n):
        no = i + 1
        f0 = f"{no:03d}_0.jpg"
        f100 = f"{no:03d}_100.jpg"
        save_jpeg(foto0[i], OUT_DIR / f0)
        zip_names.append(f0)
        rows[i]["foto_0"] = f0
        if i in assigned_100:
            save_jpeg(assigned_100[i]["img"], OUT_DIR / f100, quality=88)
            zip_names.append(f100)
            rows[i]["foto_100"] = f100
            # Update lat/lon from 100% GPS if available (lebih akurat untuk dokumentasi selesai)
            if assigned_100[i]["gps"]:
                rows[i]["lat_100"] = assigned_100[i]["gps"][0]
                rows[i]["lon_100"] = assigned_100[i]["gps"][1]
        else:
            rows[i]["foto_100"] = None

    if OUT_ZIP.exists():
        OUT_ZIP.unlink()
    with zipfile.ZipFile(OUT_ZIP, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name in zip_names:
            zf.write(OUT_DIR / name, arcname=name)

    # --- Excel ---
    wb = openpyxl.load_workbook(TEMPLATE)
    ws_info = wb["Info"]
    ws_data = wb["Data"]
    ws_petunjuk = wb["Petunjuk"]
    if ws_data.max_row > 1:
        ws_data.delete_rows(2, ws_data.max_row - 1)

    ws_info["B2"] = "Sambungan Rumah"
    ws_info["C2"] = "Unit"
    ws_info["D2"] = float(n)
    ws_info["E2"] = "unit"
    ws_info["F2"] = n

    ws_petunjuk["A1"] = "TEMPLATE IMPORT UNIT — Sambungan Rumah (RT04 + RT08 Desa Ciloto)"
    ws_petunjuk["A8"] = (
        f"Target baris: {n}. Sumber nama: DAFTAR PENERIMA RT08+RT04. "
        f"Foto 0% dari PDF daftar; foto 100% dari WhatsApp zip (match GPS watermark)."
    )
    ws_petunjuk["A15"] = (
        "Urutan: 1–29 RT.008/003, 30–55 RT.004/003. "
        "ZIP: 001_0.jpg + 001_100.jpg dst."
    )
    n100 = sum(1 for r in rows if r.get("foto_100"))
    ws_petunjuk["A16"] = f"Foto 100% ter-assign: {n100}/{n}."

    for i, row in enumerate(rows, start=1):
        # Prefer 100% GPS for import coords when available
        lat = row.get("lat_100", row["lat"])
        lon = row.get("lon_100", row["lon"])
        ws_data.append(
            [
                i,
                row["nama"],
                row["nik"],
                row["alamat"],
                row["jumlah_jiwa"],
                lat,
                lon,
                row.get("foto_0"),
                None,
                None,
                None,
                row.get("foto_100"),
            ]
        )

    wb.save(OUT_XLSX)
    try:
        wb.save(OUT_XLSX_DOCS)
    except OSError:
        pass

    size_mb = OUT_ZIP.stat().st_size / (1024 * 1024)
    print()
    print(f"Excel: {OUT_XLSX}")
    print(f"ZIP:   {OUT_ZIP} ({size_mb:.2f} MB), files={len(zip_names)}")
    print(f"Dir:   {OUT_DIR}")
    print(f"Foto 0%: {n}, Foto 100%: {n100}")


if __name__ == "__main__":
    main()
