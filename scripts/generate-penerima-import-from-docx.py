#!/usr/bin/env python3
"""Generate Arumanis penerima import Excel + foto ZIP from a Word recipient list."""

from __future__ import annotations

import argparse
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import Workbook

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
R = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

FOTO_HEADERS = [
    "nama_file_foto_0",
    "nama_file_foto_25",
    "nama_file_foto_50",
    "nama_file_foto_75",
    "nama_file_foto_100",
]
DATA_HEADERS = [
    "no",
    "nama",
    "nik",
    "alamat",
    "jumlah_jiwa",
    "latitude",
    "longitude",
    *FOTO_HEADERS,
]


def text_of(element: ET.Element) -> str:
    parts: list[str] = []
    for node in element.iter(f"{W}t"):
        if node.text:
            parts.append(node.text)
        if node.tail:
            parts.append(node.tail)
    return re.sub(r"\s+", " ", "".join(parts)).strip()


def embed_ids(cell: ET.Element) -> list[str]:
    ids: list[str] = []
    for blip in cell.iter(f"{A}blip"):
        embed = blip.attrib.get(R + "embed")
        if embed:
            ids.append(embed)
    return ids


def parse_no(raw: str) -> int | None:
    match = re.search(r"(\d+)", raw)
    return int(match.group(1)) if match else None


def parse_jumlah_jiwa(raw: str) -> int:
    match = re.search(r"(\d+)\s*JIWA", raw, re.I)
    return int(match.group(1)) if match else 1


def parse_coords(raw: str) -> tuple[float | None, float | None]:
    lat_match = re.search(r"LAT:\s*(-?\d+(?:\.\d+)?)", raw, re.I)
    long_match = re.search(r"LONG:\s*(-?\d+(?:\.\d+)?)", raw, re.I)
    lat = float(lat_match.group(1)) if lat_match else None
    lng = float(long_match.group(1)) if long_match else None
    return lat, lng


def parse_nama_alamat_nik(raw: str) -> tuple[str, str, str]:
    nik_match = re.search(r"NIK\s*:\s*(\d{16})", raw, re.I)
    nik = nik_match.group(1) if nik_match else ""
    body = raw[: nik_match.start()] if nik_match else raw
    body = re.sub(r"NIK\s*:.*$", "", body, flags=re.I).strip()

    glued = re.match(r"^([A-Za-z]+)(KP\..*)$", body, re.I | re.S)
    if glued:
        return glued.group(1).strip().upper(), re.sub(r"\s+", " ", glued.group(2)).strip(), nik

    spaced = re.match(r"^([A-Za-z][A-Za-z\s]*?)\s+(KP\..*)$", body, re.I | re.S)
    if spaced:
        return spaced.group(1).strip().upper(), re.sub(r"\s+", " ", spaced.group(2)).strip(), nik

    spaced_kp = re.match(r"^([A-Za-z][A-Za-z\s]*?)\s+(KP\s+.*)$", body, re.I | re.S)
    if spaced_kp:
        return spaced_kp.group(1).strip().upper(), re.sub(r"\s+", " ", spaced_kp.group(2)).strip(), nik

    tokens = body.split()
    if tokens:
        return tokens[0].upper(), body, nik
    return "", body, nik


def guess_extension(media_path: str) -> str:
    suffix = Path(media_path).suffix.lower()
    if suffix in {".jpg", ".jpeg", ".png", ".webp"}:
        return ".jpg" if suffix == ".jpeg" else suffix
    return ".jpg"


def extract_rows(docx_path: Path) -> tuple[list[dict], dict[str, bytes]]:
    rows: list[dict] = []
    media_files: dict[str, bytes] = {}

    with zipfile.ZipFile(docx_path) as archive:
        for name in archive.namelist():
            if name.startswith("word/media/"):
                media_files[name] = archive.read(name)

        root = ET.fromstring(archive.read("word/document.xml"))
        rels_root = ET.fromstring(archive.read("word/_rels/document.xml.rels"))
        rel_map = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels_root
            if rel.attrib.get("Id") and rel.attrib.get("Target")
        }

        for table_row in root.iter(f"{W}tr"):
            cells = table_row.findall(f"{W}tc")
            if len(cells) < 4:
                continue

            texts = [text_of(cell) for cell in cells[:5]]
            if texts[0].upper().startswith("NO") or "NAMA" in texts[1].upper():
                continue

            no = parse_no(texts[0])
            nama, alamat, nik = parse_nama_alamat_nik(texts[1])
            jumlah_jiwa = parse_jumlah_jiwa(texts[2])
            latitude, longitude = parse_coords(texts[3])
            embeds = embed_ids(cells[4]) if len(cells) >= 5 else []

            media_targets: list[str] = []
            for embed in embeds:
                target = rel_map.get(embed, "")
                if target.startswith("media/"):
                    target = f"word/{target}"
                if target in media_files:
                    media_targets.append(target)

            rows.append(
                {
                    "no": no,
                    "nama": nama,
                    "nik": nik,
                    "alamat": alamat,
                    "jumlah_jiwa": jumlah_jiwa,
                    "latitude": latitude,
                    "longitude": longitude,
                    "media_targets": media_targets,
                }
            )

    rows.sort(key=lambda item: item["no"] or 0)
    return rows, media_files


def build_workbook(rows: list[dict], source_name: str) -> Workbook:
    workbook = Workbook()
    info = workbook.active
    info.title = "Info"
    info.append(["komponen_id", "komponen", "satuan", "volume", "tipe_import", "target_baris"])
    info.append(["", source_name, "unit", len(rows), "unit", len(rows)])

    data = workbook.create_sheet("Data")
    data.append(DATA_HEADERS)

    for index, row in enumerate(rows, start=1):
        row_no = row["no"] or index
        foto_name = ""
        if row["media_targets"]:
            ext = guess_extension(row["media_targets"][0])
            foto_name = f"{row_no:03d}_0{ext}"

        data.append(
            [
                row_no,
                row["nama"],
                row["nik"],
                row["alamat"],
                row["jumlah_jiwa"],
                row["latitude"] if row["latitude"] is not None else "",
                row["longitude"] if row["longitude"] is not None else "",
                foto_name,
                "",
                "",
                "",
                "",
            ]
        )

    guide = workbook.create_sheet("Petunjuk")
    guide.append([f"IMPORT PENERIMA — {source_name}"])
    guide.append([""])
    guide.append(["File ini dibuat otomatis dari dokumen Word."])
    guide.append(["Saat import di Arumanis, pilih komponen/output pekerjaan yang sesuai."])
    guide.append(["Kosongkan atau abaikan komponen_id di sheet Info jika tidak cocok."])
    guide.append([f"Total penerima: {len(rows)}"])
    guide.append([f"Baris dengan foto 0%: {sum(1 for row in rows if row['media_targets'])}"])

    return workbook


def build_foto_zip(rows: list[dict], media_files: dict[str, bytes], zip_path: Path) -> int:
    written = 0
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for index, row in enumerate(rows, start=1):
            if not row["media_targets"]:
                continue
            row_no = row["no"] or index
            media_path = row["media_targets"][0]
            ext = guess_extension(media_path)
            filename = f"{row_no:03d}_0{ext}"
            archive.writestr(filename, media_files[media_path])
            written += 1
    return written


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("docx", type=Path)
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path.home() / "Downloads",
    )
    parser.add_argument("--slug", default="rawabelut")
    args = parser.parse_args()

    rows, media_files = extract_rows(args.docx)
    if not rows:
        raise SystemExit("Tidak ada baris penerima yang terbaca dari dokumen.")

    slug = args.slug
    source_name = f"PAM DESA {slug.upper()}"
    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    excel_path = output_dir / f"import_penerima_{slug}.xlsx"
    zip_path = output_dir / f"import_penerima_{slug}_foto.zip"

    workbook = build_workbook(rows, source_name)
    workbook.save(excel_path)
    foto_count = build_foto_zip(rows, media_files, zip_path)

    print(f"Excel: {excel_path}")
    print(f"ZIP foto: {zip_path}")
    print(f"Penerima: {len(rows)}")
    print(f"Foto 0%: {foto_count}")


if __name__ == "__main__":
    main()